# src/app.py
from flask import Flask, render_template, jsonify, request, send_from_directory, Response
from flask_socketio import SocketIO, emit
from flask_cors import CORS
import os
from pathlib import Path
import json
import tempfile
import cv2
from ultralytics import YOLO
from datetime import datetime
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill
import io
from PIL import Image as PILImage

app = Flask(__name__, static_folder="../static", template_folder="templates")
CORS(app, resources={r"/*": {"origins": "*"}})
socketio = SocketIO(app, cors_allowed_origins="*")

ROOT = Path(__file__).resolve().parents[1]

ALERTS_DIR = ROOT / "alerts"
ALERTS_DIR.mkdir(exist_ok=True)

DETECTIONS_DIR = ROOT / "detections"
DETECTIONS_DIR.mkdir(exist_ok=True)

SCREENSHOTS_DIR = ROOT / "screenshots"
SCREENSHOTS_DIR.mkdir(exist_ok=True)

EXCEL_FILE = ROOT / "detection_alerts.xlsx"

# Load models (lazy loading)
MODELS = {}
VIDEO_SESSIONS = {}  # Store active video sessions for streaming
STOP_FLAGS = {}  # Flags to stop streaming for each camera

def get_model(model_type):
    """Lazy load models"""
    if model_type not in MODELS:
        model_paths = {
            'crowd': 'A:/src1/models/crowd_yolo6/weights/best.pt',
            'weapon': 'A:/src1/models/weapon4/weights/best.pt',
            'fight': 'A:/src1/models/fight_yolo/weights/best.pt'
        }
        if model_type in model_paths and Path(model_paths[model_type]).exists():
            MODELS[model_type] = YOLO(model_paths[model_type])
            print(f"✅ Loaded {model_type} model")
    return MODELS.get(model_type)

def log_detection(camera_id, detection_type, count=0, confidence=0.0):
    """Log detection data for analytics"""
    timestamp = datetime.now()
    date_str = timestamp.strftime('%Y-%m-%d')
    
    detection_file = DETECTIONS_DIR / f"{date_str}.json"
    
    # Load existing data
    if detection_file.exists():
        with open(detection_file, 'r') as f:
            data = json.load(f)
    else:
        data = []
    
    # Add new detection
    data.append({
        'timestamp': timestamp.isoformat(),
        'camera': camera_id,
        'type': detection_type,
        'count': count,
        'confidence': confidence
    })
    
    # Save updated data
    with open(detection_file, 'w') as f:
        json.dump(data, f, indent=2)

def save_detection_to_excel(camera_id, detection_type, confidence, frame):
    """Save detection alert to Excel with screenshot"""
    try:
        timestamp = datetime.now()
        
        # Create or load Excel workbook
        if EXCEL_FILE.exists():
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Detection Alerts"
            
            # Create header row
            headers = ["#", "Date", "Time", "Camera", "Detection Type", "Confidence", "Screenshot"]
            ws.append(headers)
            
            # Style header
            header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Set column widths
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 18
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 25
        
        # Save screenshot
        screenshot_filename = f"{detection_type}_{camera_id}_{timestamp.strftime('%Y%m%d_%H%M%S')}.jpg"
        screenshot_path = SCREENSHOTS_DIR / screenshot_filename
        cv2.imwrite(str(screenshot_path), frame)
        
        # Get next row number
        row_num = ws.max_row + 1
        alert_num = row_num - 1
        
        # Add data to Excel
        date_str = timestamp.strftime('%Y-%m-%d')
        time_str = timestamp.strftime('%H:%M:%S')
        conf_str = f"{confidence:.2%}" if confidence > 0 else "N/A"
        
        ws.cell(row=row_num, column=1, value=alert_num)
        ws.cell(row=row_num, column=2, value=date_str)
        ws.cell(row=row_num, column=3, value=time_str)
        ws.cell(row=row_num, column=4, value=camera_id)
        ws.cell(row=row_num, column=5, value=detection_type.upper())
        ws.cell(row=row_num, column=6, value=conf_str)
        ws.cell(row=row_num, column=7, value=screenshot_filename)
        
        # Center align all cells in this row
        for col_num in range(1, 8):
            ws.cell(row=row_num, column=col_num).alignment = Alignment(horizontal='center', vertical='center')
        
        # Color code by detection type
        type_colors = {
            'weapon': 'FFCCCC',  # Light red
            'fight': 'FFE6CC',   # Light orange
            'crowd': 'FFFFCC'    # Light yellow
        }
        
        if detection_type in type_colors:
            type_fill = PatternFill(start_color=type_colors[detection_type], 
                                   end_color=type_colors[detection_type], 
                                   fill_type="solid")
            for col_num in range(1, 8):
                ws.cell(row=row_num, column=col_num).fill = type_fill
        
        # Set row height for better visibility
        ws.row_dimensions[row_num].height = 20
        
        # Save workbook
        wb.save(EXCEL_FILE)
        print(f"✅ Saved {detection_type} detection to Excel: Row {alert_num}")
        
    except Exception as e:
        print(f"❌ Error saving to Excel: {e}")

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/two_cam")
def two_cam():
    return render_template("two_cam.html")

@app.route("/api/alerts", methods=["POST"])
def receive_alert():
    data = request.get_json() or request.form.to_dict()
    
    # Parse JSON if it's in a string field
    if "json" in data:
        try:
            data = json.loads(data["json"])
        except:
            pass
    
    fname = (ALERTS_DIR / f"alert_{len(list(ALERTS_DIR.glob('*.json')))+1}.json")
    with open(fname, "w") as f:
        json.dump(data, f, indent=2)
    print("Received alert:", fname)
    
    # Broadcast alert to WebSocket clients
    socketio.emit('new_alert', data)
    print(f"📡 Broadcasted alert via WebSocket: {data.get('type', 'unknown')}")
    
    return jsonify({"status": "ok", "saved": str(fname)})

@app.route("/api/alerts/list")
def list_alerts():
    files = sorted([p.name for p in ALERTS_DIR.glob("*.json")], reverse=True)
    return jsonify(files)

@app.route("/api/alerts/get/<fname>")
def get_alert(fname):
    p = ALERTS_DIR / fname
    if p.exists():
        return p.read_text(), 200, {"Content-Type": "application/json"}
    return jsonify({"error": "not found"}), 404

@app.route("/api/start_live_camera/<camera_id>", methods=["POST"])
def start_live_camera(camera_id):
    """Start live camera feed"""
    try:
        data = request.get_json() or {}
        camera_index = data.get('camera_index', 0)  # Default to camera 0
        
        print(f"📹 Starting live camera {camera_index} for {camera_id}")
        
        # Stop old stream if exists
        if camera_id in VIDEO_SESSIONS:
            STOP_FLAGS[camera_id] = True
            import time
            time.sleep(0.5)
        
        # Store camera index as negative number to distinguish from video files
        VIDEO_SESSIONS[camera_id] = f"camera:{camera_index}"
        STOP_FLAGS[camera_id] = False
        
        return jsonify({
            "status": "success",
            "camera_id": camera_id,
            "stream_url": f"/api/video_feed/{camera_id}",
            "camera_index": camera_index
        })
        
    except Exception as e:
        print(f"❌ Error starting live camera: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/detect", methods=["POST"])
def detect_video():
    """Process uploaded video and stream frames with detection to frontend"""
    try:
        print("📥 Received video upload request")
        
        if 'video' not in request.files:
            print("❌ No video file in request")
            return jsonify({"error": "No video file provided"}), 400
        
        video_file = request.files['video']
        camera_id = request.form.get('camera_id', 'cam-1')
        print(f"📹 Processing video: {video_file.filename} for camera: {camera_id}")
        
        # Save temporarily
        with tempfile.NamedTemporaryFile(delete=False, suffix='.mp4') as tmp:
            video_file.save(tmp.name)
            tmp_path = tmp.name
        
        print(f"💾 Saved to: {tmp_path}")
        
        # Stop old video stream if exists
        if camera_id in VIDEO_SESSIONS:
            STOP_FLAGS[camera_id] = True
            print(f"🛑 Stopping old stream for {camera_id}")
            import time
            time.sleep(0.5)  # Wait for old stream to stop
        
        # Clean up old video if exists
        old_path = VIDEO_SESSIONS.get(camera_id)
        if old_path and os.path.exists(old_path) and old_path != tmp_path:
            try:
                os.unlink(old_path)
                print(f"🗑️ Deleted old video: {old_path}")
            except Exception as e:
                print(f"⚠️ Could not delete old video: {e}")
        
        # Store video path for streaming and reset stop flag
        VIDEO_SESSIONS[camera_id] = tmp_path
        STOP_FLAGS[camera_id] = False
        
        return jsonify({
            "status": "success",
            "camera_id": camera_id,
            "stream_url": f"/api/video_feed/{camera_id}"
        })
        
    except Exception as e:
        print(f"❌ Error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/video_feed/<camera_id>")
def video_feed(camera_id):
    """Stream video frames with detection overlays"""
    def generate():
        video_path = VIDEO_SESSIONS.get(camera_id)
        if not video_path:
            return
        
        # Check if it's a live camera or video file
        is_live_camera = isinstance(video_path, str) and video_path.startswith("camera:")
        
        if is_live_camera:
            # Extract camera index
            camera_index = int(video_path.split(":")[1])
            cap = cv2.VideoCapture(camera_index, cv2.CAP_DSHOW)  # Use DirectShow for Windows
            if not cap.isOpened():
                print(f"❌ Failed to open camera {camera_index}, trying without CAP_DSHOW")
                cap = cv2.VideoCapture(camera_index)
            
            if not cap.isOpened():
                print(f"❌ Failed to open camera {camera_index}")
                return
            
            # Set camera properties for better performance
            cap.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
            cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
            cap.set(cv2.CAP_PROP_FPS, 30)
            
            print(f"📹 Starting live camera stream: {camera_index}")
        else:
            # Video file
            if not os.path.exists(video_path):
                return
            cap = cv2.VideoCapture(video_path)
            print(f"📹 Starting video file stream: {video_path}")
        crowd_model = get_model('crowd')
        weapon_model = get_model('weapon')
        fight_model = get_model('fight')
        
        frame_count = 0
        weapon_detected = False
        weapon_conf = 0.0
        fight_detected = False
        fight_conf = 0.0
        
        # Persistent detection state for continuous display
        last_person_boxes = []
        last_weapon_boxes = []
        current_count = 0
        
        # Group detection tracking (for 5+ people alert)
        group_start_time = None
        group_alert_sent = False
        fps = cap.get(cv2.CAP_PROP_FPS) or 30  # Get video FPS
        
        # Fight detection tracking (require sustained detection)
        fight_frame_count = 0
        fight_threshold_frames = int(fps * 3)  # Require 3 seconds of sustained fight detection
        fight_alert_sent = False
        last_fight_time = 0
        
        # Performance optimization - track real time
        import time
        start_time = time.time()
        last_frame_time = start_time
        
        while cap.isOpened():
            # Check if we should stop this stream
            if STOP_FLAGS.get(camera_id, False):
                print(f"🛑 Stream stopped for {camera_id}")
                break
            
            ret, frame = cap.read()
            if not ret:
                # For live camera, break on error
                if is_live_camera:
                    print(f"❌ Live camera read error for {camera_id}")
                    break
                # For video file, loop - restart from beginning
                cap.set(cv2.CAP_PROP_POS_FRAMES, 0)
                frame_count = 0
                continue
            
            frame_count += 1
            display_frame = frame.copy()
            
            # Calculate real elapsed time for accurate timing
            current_time = time.time()
            elapsed_time = current_time - start_time
            
            # Optimize detection - run crowd/fight every frame, weapon every 2 frames
            run_weapon_detection = (frame_count % 2 == 0)
            
            # Run detection on EVERY frame for continuous detection (not just every 5th)
            # Detect people on every frame with optimized settings for crowds
            if crowd_model:
                results = crowd_model.predict(frame, conf=0.15, verbose=False, imgsz=640, iou=0.4, max_det=100)
                for r in results:
                    if hasattr(r, 'boxes'):
                        boxes = r.boxes.xyxy.cpu().numpy()
                        current_count = len(boxes)
                        last_person_boxes = boxes  # Store for continuous display
                        
                        # Draw all detected people with green boxes
                        # If 5+ people (group), draw yellow boxes instead
                        box_color = (0, 255, 255) if current_count >= 5 else (0, 255, 0)  # Yellow for groups, green for individuals
                        box_thickness = 3 if current_count >= 5 else 2
                        
                        for box in boxes:
                            x1, y1, x2, y2 = map(int, box)
                            cv2.rectangle(display_frame, (x1, y1), (x2, y2), box_color, box_thickness)
                            # Add person label
                            cv2.putText(display_frame, "Person", (x1, y1-5),
                                      cv2.FONT_HERSHEY_SIMPLEX, 0.6, box_color, 2)
                        
                        # If group detected, show timer on screen near the group
                        if current_count >= 5 and len(boxes) > 0:
                            # Calculate center of all boxes for timer placement
                            center_x = int(sum([box[0] + box[2] for box in boxes]) / (2 * len(boxes)))
                            center_y = int(sum([box[1] + box[3] for box in boxes]) / (2 * len(boxes)))
                            
                            # Calculate timer duration using REAL TIME
                            if group_start_time is not None:
                                group_duration = int(elapsed_time - group_start_time)
                                timer_text = f"GROUP: {group_duration}s"
                                timer_color = (0, 255, 255) if group_duration < 60 else (0, 0, 255)
                                
                                # Draw timer with background
                                timer_size = cv2.getTextSize(timer_text, cv2.FONT_HERSHEY_SIMPLEX, 1.2, 3)[0]
                                timer_x = max(10, center_x - timer_size[0] // 2)
                                timer_y = max(50, center_y - 30)
                                
                                # Background rectangle for timer
                                cv2.rectangle(display_frame, 
                                            (timer_x - 10, timer_y - timer_size[1] - 10),
                                            (timer_x + timer_size[0] + 10, timer_y + 10),
                                            (0, 0, 0), -1)
                                cv2.rectangle(display_frame, 
                                            (timer_x - 10, timer_y - timer_size[1] - 10),
                                            (timer_x + timer_size[0] + 10, timer_y + 10),
                                            timer_color, 2)
                                
                                # Timer text (large and bold)
                                cv2.putText(display_frame, timer_text, (timer_x, timer_y),
                                          cv2.FONT_HERSHEY_SIMPLEX, 1.2, timer_color, 3)
                        
                        # Check for group (5+ people)
                        if current_count >= 5:
                            if group_start_time is None:
                                group_start_time = elapsed_time
                                print(f"👥 Group detected: {current_count} people")
                            else:
                                # Calculate how long group has been present using REAL TIME
                                group_duration = elapsed_time - group_start_time
                                
                                # Send alert after 60 seconds (1 minute)
                                if group_duration >= 60 and not group_alert_sent:
                                    socketio.emit('new_alert', {
                                        'type': 'crowd',
                                        'camera': camera_id,
                                        'count': current_count,
                                        'duration': int(group_duration),
                                        'timestamp': datetime.now().isoformat(),
                                        'severity': 'medium',
                                        'message': f'{current_count} people detected for {int(group_duration)} seconds'
                                    })
                                    # Save to Excel with screenshot
                                    save_detection_to_excel(camera_id, 'crowd', 0.0, display_frame.copy())
                                    # Log to JSON for analytics
                                    log_detection(camera_id, 'crowd', count=current_count, confidence=0.0)
                                    group_alert_sent = True
                                    print(f"🚨 Group alert sent: {current_count} people for {int(group_duration)}s")
                        else:
                            # Reset group tracking if count drops below 5
                            group_start_time = None
                            group_alert_sent = False
            
            # Detect weapons every 2 frames (optimization for smooth playback)
            weapon_detected = False
            if weapon_model and run_weapon_detection:
                results = weapon_model.predict(frame, conf=0.2, verbose=False, imgsz=640)
                for r in results:
                    if hasattr(r, 'boxes') and len(r.boxes) > 0:
                        boxes = r.boxes.xyxy.cpu().numpy()
                        confs = r.boxes.conf.cpu().numpy()
                        weapon_detected = True
                        weapon_conf = max(confs) if len(confs) > 0 else 0.0
                        last_weapon_boxes = list(zip(boxes, confs))
                        
                        # Send weapon alert via WebSocket (throttle to avoid spam)
                        if frame_count % 30 == 0:
                            socketio.emit('new_alert', {
                                'type': 'weapon',
                                'camera': camera_id,
                                'confidence': float(weapon_conf),
                                'timestamp': datetime.now().isoformat(),
                                'severity': 'high'
                            })
                            # Save to Excel with screenshot
                            save_detection_to_excel(camera_id, 'weapon', weapon_conf, display_frame.copy())
                            # Log to JSON for analytics
                            log_detection(camera_id, 'weapon', count=len(boxes), confidence=float(weapon_conf))
                        
                        # Draw all weapon boxes with red (large and visible)
                        for box, conf in zip(boxes, confs):
                            x1, y1, x2, y2 = map(int, box)
                            # Thick red box for weapon
                            cv2.rectangle(display_frame, (x1, y1), (x2, y2), (0, 0, 255), 4)
                            
                            # Large weapon label with background
                            weapon_label = f"WEAPON {conf:.2f}"
                            label_size = cv2.getTextSize(weapon_label, cv2.FONT_HERSHEY_SIMPLEX, 1.0, 3)[0]
                            label_y = max(30, y1 - 10)
                            
                            # Black background for better visibility
                            cv2.rectangle(display_frame, 
                                        (x1, label_y - label_size[1] - 5),
                                        (x1 + label_size[0] + 5, label_y + 5),
                                        (0, 0, 0), -1)
                            # Red border around label
                            cv2.rectangle(display_frame, 
                                        (x1, label_y - label_size[1] - 5),
                                        (x1 + label_size[0] + 5, label_y + 5),
                                        (0, 0, 255), 2)
                            
                            # Large weapon text
                            cv2.putText(display_frame, weapon_label, (x1 + 2, label_y), 
                                      cv2.FONT_HERSHEY_SIMPLEX, 1.0, (0, 0, 255), 3)
                        
                        # Log weapon detection
                        if frame_count % 30 == 0:
                            log_detection(camera_id, 'weapon', count=len(boxes), confidence=float(weapon_conf))
            
            # Detect fights on every frame (with stricter requirements)
            fight_detected = False
            if fight_model:
                results = fight_model.predict(frame, conf=0.65, verbose=False, imgsz=640)  # Increased threshold to 0.65
                for r in results:
                    if hasattr(r, 'boxes') and len(r.boxes) > 0:
                        boxes = r.boxes.xyxy.cpu().numpy()
                        confs = r.boxes.conf.cpu().numpy()
                        max_conf = max(confs) if len(confs) > 0 else 0.0
                        
                        # Only count if confidence is high (0.65+)
                        if max_conf >= 0.65:
                            # Check if detection is continuous (within 1 second of last detection)
                            if frame_count - last_fight_time <= fps:
                                fight_frame_count += 1
                            else:
                                # Reset if gap is too large
                                fight_frame_count = 1
                            
                            last_fight_time = frame_count
                            
                            # Only mark as detected if sustained over threshold (3 seconds)
                            if fight_frame_count >= fight_threshold_frames:
                                fight_detected = True
                                fight_conf = max_conf
                                
                                # Send fight alert once (throttled)
                                if not fight_alert_sent:
                                    socketio.emit('new_alert', {
                                        'type': 'fight',
                                        'camera': camera_id,
                                        'confidence': float(fight_conf),
                                        'timestamp': datetime.now().isoformat(),
                                        'severity': 'high'
                                    })
                                    # Save to Excel with screenshot
                                    save_detection_to_excel(camera_id, 'fight', fight_conf, display_frame.copy())
                                    fight_alert_sent = True
                                    log_detection(camera_id, 'fight', count=len(boxes), confidence=float(fight_conf))
                            
                            # Draw all fight boxes with orange (only if sustained)
                            if fight_detected:
                                for box, conf in zip(boxes, confs):
                                    x1, y1, x2, y2 = map(int, box)
                                    # Thick orange box for fight
                                    cv2.rectangle(display_frame, (x1, y1), (x2, y2), (0, 140, 255), 4)
                                    
                                    # Large fight label with background
                                    fight_label = f"FIGHT {conf:.2f}"
                                    label_size = cv2.getTextSize(fight_label, cv2.FONT_HERSHEY_SIMPLEX, 1.0, 3)[0]
                                    label_y = max(30, y1 - 10)
                                    
                                    # Black background for better visibility
                                    cv2.rectangle(display_frame, 
                                                (x1, label_y - label_size[1] - 5),
                                                (x1 + label_size[0] + 5, label_y + 5),
                                                (0, 0, 0), -1)
                                    # Orange border around label
                                    cv2.rectangle(display_frame, 
                                                (x1, label_y - label_size[1] - 5),
                                                (x1 + label_size[0] + 5, label_y + 5),
                                                (0, 140, 255), 2)
                                    
                                    # Large fight text
                                    cv2.putText(display_frame, fight_label, (x1 + 2, label_y), 
                                              cv2.FONT_HERSHEY_SIMPLEX, 1.0, (0, 140, 255), 3)
                        else:
                            # Low confidence - reset counter if gap exceeds 1 second
                            if frame_count - last_fight_time > fps:
                                fight_frame_count = 0
                                fight_alert_sent = False
                    else:
                        # No detection - reset counter if gap exceeds 1 second
                        if frame_count - last_fight_time > fps:
                            fight_frame_count = 0
                            fight_alert_sent = False
            
            # Draw HUD with larger text
            hud_x, hud_y = 10, 10
            overlay = display_frame.copy()
            cv2.rectangle(overlay, (hud_x, hud_y), (hud_x + 350, hud_y + 130), (0, 0, 0), -1)
            cv2.addWeighted(overlay, 0.5, display_frame, 0.5, 0, display_frame)
            
            tx, ty = hud_x + 8, hud_y + 30
            weapon_color = (0, 0, 255) if weapon_detected else (0, 255, 0)
            weapon_text = f"Weapon: {'DETECTED' if weapon_detected else 'SAFE'}"
            cv2.putText(display_frame, weapon_text, (tx, ty), cv2.FONT_HERSHEY_SIMPLEX, 0.7, weapon_color, 2)
            ty += 30
            
            # Fight status
            fight_color = (0, 140, 255) if fight_detected else (0, 255, 0)
            fight_text = f"Fight: {'DETECTED' if fight_detected else 'SAFE'}"
            cv2.putText(display_frame, fight_text, (tx, ty), cv2.FONT_HERSHEY_SIMPLEX, 0.7, fight_color, 2)
            ty += 30
            
            # People count with color coding
            people_color = (0, 255, 255) if current_count >= 5 else (255, 255, 255)
            cv2.putText(display_frame, f"People: {current_count}", (tx, ty), cv2.FONT_HERSHEY_SIMPLEX, 0.7, people_color, 2)
            
            # Show group timing if 5+ people detected
            if current_count >= 5 and group_start_time is not None:
                ty += 30
                group_duration = int(elapsed_time - group_start_time)
                group_color = (0, 255, 255) if group_duration < 60 else (0, 0, 255)
                cv2.putText(display_frame, f"Group: {group_duration}s", (tx, ty), cv2.FONT_HERSHEY_SIMPLEX, 0.7, group_color, 2)
            
            # Add small delay for smooth playback (target 30 FPS)
            target_frame_time = 1.0 / 30.0  # 33ms per frame
            actual_frame_time = time.time() - current_time
            if actual_frame_time < target_frame_time:
                time.sleep(target_frame_time - actual_frame_time)
            
            # Encode frame as JPEG
            ret, buffer = cv2.imencode('.jpg', display_frame)
            frame_bytes = buffer.tobytes()
            
            yield (b'--frame\r\n'
                   b'Content-Type: image/jpeg\r\n\r\n' + frame_bytes + b'\r\n')
        
        cap.release()
        # Don't delete here - video will be cleaned up when new video is uploaded
        print(f"📹 Stream ended for {camera_id}")
    
    return Response(generate(), mimetype='multipart/x-mixed-replace; boundary=frame')

@app.route("/api/stop_video/<camera_id>", methods=["POST"])
def stop_video(camera_id):
    """Stop video stream and cleanup"""
    try:
        # Set stop flag to terminate stream
        STOP_FLAGS[camera_id] = True
        
        import time
        time.sleep(0.3)  # Wait for stream to stop
        
        video_path = VIDEO_SESSIONS.get(camera_id)
        if video_path and os.path.exists(video_path):
            os.unlink(video_path)
            print(f"🗑️ Deleted video for {camera_id}")
        
        if camera_id in VIDEO_SESSIONS:
            del VIDEO_SESSIONS[camera_id]
        if camera_id in STOP_FLAGS:
            del STOP_FLAGS[camera_id]
        
        return jsonify({"status": "success"})
    except Exception as e:
        print(f"⚠️ Error stopping video: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/detections/<date>", methods=["GET"])
def get_detections(date):
    """Get detection data for a specific date"""
    try:
        detection_file = DETECTIONS_DIR / f"{date}.json"
        if not detection_file.exists():
            return jsonify([])
        
        with open(detection_file, 'r') as f:
            data = json.load(f)
        
        # Filter by type if specified
        filter_type = request.args.get('type', None)
        if filter_type:
            data = [d for d in data if d['type'] == filter_type]
        
        return jsonify(data)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/detections/range", methods=["GET"])
def get_detections_range():
    """Get detection data for a date range"""
    try:
        start_date = request.args.get('start')
        end_date = request.args.get('end')
        filter_type = request.args.get('type', None)
        
        all_data = []
        
        # Get all detection files
        for file in sorted(DETECTIONS_DIR.glob("*.json")):
            file_date = file.stem
            if (not start_date or file_date >= start_date) and (not end_date or file_date <= end_date):
                with open(file, 'r') as f:
                    data = json.load(f)
                    if filter_type:
                        data = [d for d in data if d['type'] == filter_type]
                    all_data.extend(data)
        
        return jsonify(all_data)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/open_powerbi", methods=["POST"])
def open_powerbi():
    """Open Power BI dashboard"""
    try:
        data = request.get_json()
        powerbi_path = data.get('path', 'A:\\Manav Mule power bi.pbix')
        
        if not os.path.exists(powerbi_path):
            return jsonify({"error": f"Power BI file not found: {powerbi_path}"}), 404
        
        # Open Power BI file using Windows default application
        import subprocess
        subprocess.Popen(['start', '', powerbi_path], shell=True)
        
        print(f"📊 Opened Power BI dashboard: {powerbi_path}")
        return jsonify({"status": "success", "message": "Power BI dashboard opened"})
        
    except Exception as e:
        print(f"❌ Error opening Power BI: {e}")
        return jsonify({"error": str(e)}), 500

@socketio.on('connect')
def handle_connect():
    print('Client connected')
    emit('connection_status', {'status': 'connected'})

@socketio.on('disconnect')
def handle_disconnect():
    print('Client disconnected')

if __name__ == "__main__":
    # flask-socketio run
    socketio.run(app, host="0.0.0.0", port=5000, debug=True)
